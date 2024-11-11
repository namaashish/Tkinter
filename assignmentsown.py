from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import tkinter as tk
import openpyxl 
from openpyxl import Workbook
import pathlib

#part1
root = Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False,False)


file=pathlib.Path('backend_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "PhoneNumber"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    file.save('backend_data.xlsx')





def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)

    


      
    
    if not name or not contact or not age or not gender or not address :
        messagebox.showerror('Input error','Enter details in all boxes')
        return
       

    

   

    file=openpyxl.load_workbook("backend_data.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    existing_values = [cell.value for cell in sheet[2] if cell.value is not NONE ]
    if contact not in existing_values:
        file.save(r'backend_data.xlsx')
        messagebox.showinfo('info','detail added!')
        
        nameValue.set('')
        contactValue.set('')
        AgeValue.set('')
        addressEntry.delete(1.0,END)
        gender_combobox.set('')
    else:
        messagebox.showinfo('info','Mobile number already present, provide a different number')






def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)
    gender_combobox.set('')





#part2heading
Label(root,text="Please fill out this entry from :", font="arail 15").place(x=20,y=20)

#part3
Label(root,text="Name :", font="arail 13").place(x=50,y=100)
Label(root,text="Contact number :",font="arail 13").place(x=50,y=150)
Label(root,text="Age :", font="arail 13").place(x=50,y=200)
Label(root,text="Gender :", font="arail 13").place(x=370,y=200)
Label(root,text="Address :", font="arail 13").place(x=50,y=250)

#part4entry
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameentry = Entry(root,textvariable=nameValue,width=45,bd=2,font=20)
contactentry = Entry(root,textvariable=contactValue,width=45,bd=2,font=20)
Ageentry = Entry(root,textvariable=AgeValue,width=15,bd=2,font=20)

#genderusingcombobox
gender_combobox = Combobox(root,values=['Male','Female','others'],font = 'arial 14',state = 'readonly',width = 14)
gender_combobox.place(x=440,y=200)
gender_combobox.set('')

addressEntry = Text(root,width=50,height=4,bd=2)


nameentry.place(x=200,y=100)
contactentry.place(x=200,y=150)
Ageentry.place(x=200,y=200)
addressEntry.place(x=200,y=250)

Button(root,text="Submit",width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text="clear",width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text="Exit",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)



root.mainloop()


